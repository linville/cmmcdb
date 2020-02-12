# This exports an OpenXML spreadsheet of the CMMC Model Framework.

from django.core.management.base import BaseCommand, CommandError
from django.db.models import Max

from cmmcdb.models import *

import argparse
import json
import datetime

from openpyxl import Workbook
from openpyxl.utils import get_column_letter
from openpyxl.styles import NamedStyle, Alignment, Font, Color, PatternFill


class Command(BaseCommand):
    """ Exports the CMMC database structure as Open XML spreadsheet."""

    help = "Exports the CMMC database structure as Open XML spreadsheet."

    def add_arguments(self, parser):
        parser.add_argument(
            "FILE",
            nargs="?",
            type=argparse.FileType("wb"),
            default="cmmc.xlsx",
            help="Path to the CMMC spreadsheet to create.",
        )

    def handle(self, *args, **options):
        self.wb = Workbook()
        self.define_styles()
        self.wb_row = 1
        # wb.active

        self.wb.freeze_panes = "A1"
        c = self.wb.active.cell(column=1, row=self.wb_row, value=f"Capability")
        c.style = "Maturity Level"

        self.ml = {}
        for ml in MaturityLevel.objects.order_by("level").all():
            self.ml[ml.level] = ml
            c = self.wb.active.cell(
                column=ml.level + 1, row=self.wb_row, value=f"Level {ml.level}"
            )
            c.style = "Maturity Level"
        self.wb_row += 1

        # Column Formatting
        self.wb.active.column_dimensions[get_column_letter(1)].width = 45
        for i in range(2, 7):
            self.wb.active.column_dimensions[get_column_letter(i)].width = 40

        # Output Domain
        for domain in Domain.objects.order_by("short").all():
            print(f"Domain: {domain.name}")
            self.wb.active.merge_cells(
                start_row=self.wb_row, start_column=1, end_row=self.wb_row, end_column=6
            )
            c = self.wb.active.cell(
                column=1, row=self.wb_row, value=f"Domain: {domain.name}"
            )
            c.style = "Domain"

            self.wb_row += 1
            for capability in domain.capabilities.order_by("process", "index").all():
                c = self.wb.active.cell(
                    column=1,
                    row=self.wb_row,
                    value=f"C{capability.index} {capability.name}",
                )
                c.style = "Practice"

                c

                if capability.process:
                    c.value = f"Process: {capability.name}"
                    c.style = "Process"

                self.export_capability(capability)
            self.wb_row += 1

        self.wb.save(filename=options["FILE"].name)

    def export_capability(self, capability):
        max_count = 0
        capability_practices = []

        for i in range(1, 6):
            column = []

            practices = capability.practices.filter(
                maturity_level=self.ml[i]
            ).order_by("practice_number").all()

            for practice in practices:
                column.append(practice)

            capability_practices.append(column)
            max_count = max(max_count, len(column))

        for i in range(0, max_count):
            self.export_capability_row(capability_practices, i)
            self.wb_row += 1

    def export_capability_row(self, capability_practices, row):
        ws = self.wb.active

        for i in range(0, 5):
            ml_data = capability_practices[i]
            
            if not row < len(ml_data):
                continue

            practice = ml_data[row]

            c = self.wb.active.cell(
                column=i + 1,
                row=self.wb_row,
                value=f"{practice.practice_number} {practice.name}",
            )
            c.style = "Practice"

    def define_styles(self):
        # Domain Cell Styles
        s = NamedStyle(name="Domain")
        s.font = Font(size=16, bold=True)
        s.fill = PatternFill("solid", fgColor="fff1ce")
        self.wb.add_named_style(s)

        # Maturity Level Styles
        s = NamedStyle(name="Maturity Level")
        s.font = Font(size=12, bold=True, color="ffffff")
        s.fill = PatternFill("solid", fgColor="0000ff")
        s.alignment = Alignment(horizontal="center")
        self.wb.add_named_style(s)

        # Practice
        s = NamedStyle(name="Practice")
        s.font = Font(size=12, bold=True)
        s.alignment = Alignment(horizontal="general", vertical="top", wrap_text=True)
        self.wb.add_named_style(s)

        # Process
        s = NamedStyle(name="Process")
        s.font = Font(size=12, bold=True)
        s.fill = PatternFill("solid", fgColor="ddddff")
        s.alignment = Alignment(horizontal="general", vertical="top", wrap_text=True)
        self.wb.add_named_style(s)
